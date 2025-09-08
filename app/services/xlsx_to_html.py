
from io import BytesIO
from openpyxl import load_workbook

def _css_color(rgb):
    if not rgb: return None
    rgb = rgb.replace('#','')
    if len(rgb)==8: rgb = rgb[2:]  # strip alpha if present
    return f"#{rgb.lower()}"


def convert_xlsx_to_html(xlsx_bytes: bytes):
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    # Column widths (approximate -> px)
    col_widths = []
    for col in range(1, ws.max_column+1):
        dim = ws.column_dimensions.get(chr(64+col))
        width = getattr(dim, 'width', 10) or 10
        # Excel width to px rough conversion
        px = int((width + 0.72) * 7)
        col_widths.append(px)

    html = ["<table class='sheet'>", "<colgroup>"]
    for w in col_widths:
        html.append(f"<col style='width:{w}px'>")
    html.append("</colgroup>")

    merges = { }
    for mr in ws.merged_cells.ranges:
        key = (mr.min_row, mr.min_col)
        merges[key] = (mr.size['rows'], mr.size['columns'])

    for r in range(1, ws.max_row+1):
        html.append("<tr>")
        c = 1
        while c <= ws.max_column:
            cell = ws.cell(r, c)
            # Skip cells covered by a merge (not the top-left)
            skip = False
            for (rr, cc), (rs, cs) in merges.items():
                if rr <= r < rr+rs and cc <= c < cc+cs and not (rr==r and cc==c):
                    skip = True
                    break
            if skip:
                c += 1
                continue

            rowspan, colspan = 1, 1
            if (r, c) in merges:
                rowspan, colspan = merges[(r, c)]

            style_parts = []
            if cell.fill and getattr(cell.fill, 'fgColor', None):
                col = _css_color(cell.fill.fgColor.rgb) if getattr(cell.fill.fgColor, 'rgb', None) else None
                if col: style_parts.append(f"background:{col}")
            if cell.font:
                if cell.font.color and getattr(cell.font.color, 'rgb', None):
                    style_parts.append(f"color:{_css_color(cell.font.color.rgb)}")
                if cell.font.bold:
                    style_parts.append("font-weight:bold")
                if cell.font.size:
                    style_parts.append(f"font-size:{int(cell.font.size)}px")
                if cell.font.name:
                    style_parts.append(f"font-family:'{cell.font.name}'")
            if cell.alignment:
                if cell.alignment.horizontal:
                    style_parts.append(f"text-align:{cell.alignment.horizontal}")
                if cell.alignment.vertical:
                    # map to CSS
                    va = {'center':'middle'}.get(cell.alignment.vertical, cell.alignment.vertical)
                    style_parts.append(f"vertical-align:{va}")
                if cell.alignment.wrap_text:
                    style_parts.append("white-space:normal")
            # Borders (simplified)
            if cell.border:
                for edge in ['left','right','top','bottom']:
                    side = getattr(cell.border, edge)
                    if side and side.style:
                        style_parts.append(f"border-{edge}:1px solid #000")

            txt = cell.value if cell.value is not None else ""
            attrs = []
            if rowspan>1: attrs.append(f"rowspan='{rowspan}'")
            if colspan>1: attrs.append(f"colspan='{colspan}'")
            style = ";".join(style_parts)
            attrs.append(f"style=\"{style}\"")
            html.append(f"<td {' '.join(attrs)}>{txt}</td>")
            c += colspan
        html.append("</tr>")
    html.append("</table>")

    css = ".sheet{border-collapse:collapse} .sheet td{padding:4px;border:1px solid #ddd}" 
    warnings = []
    assets = []
    return "\n".join(html), css, warnings, assets
